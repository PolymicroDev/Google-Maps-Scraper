from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from time import sleep
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains
from selenium.webdriver.common.actions.wheel_input import ScrollOrigin
import openpyxl

class Lead:
    def __init__(self,name,phone_number,website,rating,reviews):
        self.name = name
        self.phone_number = phone_number
        self.website = website
        self.rating = rating
        self.reviews = reviews
        
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument(f"--lang=en-uk")
driver = webdriver.Chrome(options=chrome_options)
driver.get("https://www.google.com/maps")
driver.execute_cdp_cmd("Network.setUserAgentOverride", {"userAgent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.102 Safari/537.36"})

def accept_conditions():
    accept_button = driver.find_element(By.XPATH,'//*[@id="yDmH0d"]/c-wiz/div/div/div/div[2]/div[1]/div[3]/div[1]/div[1]/form[2]/div/div/button/span')
    accept_button.click()


def search_query(query):
    place = driver.find_element(By.CLASS_NAME,"searchboxinput")
    place.send_keys(query)
    place.send_keys(Keys.ENTER)
    sleep(10)

def list_leads():
    entries = driver.find_elements(By.CLASS_NAME, "hfpxzc")
    sleep(4)
    leads = []
    for business in entries:
        try:
            business.click()
            sleep(2)
            business_name_element = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="QA0Szd"]/div/div/div[1]/div[3]/div/div[1]/div/div/div[2]/div[2]/div/div[1]/div[1]/h1')))
            business_name = business_name_element.text
            
            details = driver.find_elements(By.CLASS_NAME, "CsEnBe")

            try:
                rating_element = driver.find_element(By.XPATH, '//*[@id="QA0Szd"]/div/div/div[1]/div[3]/div/div[1]/div/div/div[2]/div[2]/div/div[1]/div[2]/div/div[1]/div[2]/span[1]/span[1]')
                rating = rating_element.text
            except:
                rating = ""
            
            try:
                reviews_element = driver.find_element(By.XPATH, '//*[@id="QA0Szd"]/div/div/div[1]/div[3]/div/div[1]/div/div/div[2]/div[2]/div/div[1]/div[2]/div/div[1]/div[2]/span[2]/span/span')
                reviews = reviews_element.text[1:len(reviews_element.text)-1]
            except:
                reviews = ""

            lead_info = Lead(business_name,"","",rating,reviews)

            for detail in details:
                if "+" in detail.text and detail.text[4].isdigit():
                    lead_info.phone_number = detail.text
                elif "." in detail.text:
                    lead_info.website = detail.text
                else:
                    continue
            leads.append(lead_info)

        except:
            print(f'Element {business} not found')
    return leads
        
        
def create_excelsheet():
    leads = list_leads()
    for lead in leads:
        print(lead.name)
        print(lead.phone_number)
        print(lead.website)
        print(lead.rating)
        print(lead.reviews)
        print("----------------------------------")

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet.cell(row=1, column=1, value="Name")
    sheet.cell(row=1, column=2, value="Phone Number")
    sheet.cell(row=1, column=3, value="Website")
    sheet.cell(row=1, column=4, value="Rating")
    sheet.cell(row=1, column=5, value="Reviews")
    sheet.cell(row=1, column=6, value="FB ads")

    for row, lead in enumerate(leads, start=2):
        sheet.cell(row=row, column=1, value=lead.name)
        sheet.cell(row=row, column=2, value=lead.phone_number)
        sheet.cell(row=row, column=3, value=lead.website)
        sheet.cell(row=row, column=4, value=lead.rating)
        sheet.cell(row=row, column=5, value=lead.reviews)

    workbook.save("New Delhi, real-estate2.xlsx")


accept_conditions()
search_query("New Delhi, real-estate")
create_excelsheet()



